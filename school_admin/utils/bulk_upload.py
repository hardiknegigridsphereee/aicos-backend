"""
Utilities for parsing bulk-upload files (CSV / XLSX) into plain row dicts
that can be fed straight into the existing onboarding serializers.

Requires `openpyxl` for .xlsx support:
    pip install openpyxl
"""
import csv
import io

from openpyxl import load_workbook


class FileParseError(Exception):
    """Raised when the uploaded file can't be parsed into rows."""
    pass


def parse_uploaded_file(uploaded_file):
    """
    Parse a CSV or XLSX Django UploadedFile into a list of dicts.

    Header names are normalized (lowercased, trimmed, spaces -> underscores)
    so 'First Name' / 'first name' / 'first_name' all map to 'first_name'.
    Blank rows are skipped.
    """
    name = (uploaded_file.name or '').lower()

    if name.endswith('.csv'):
        return _parse_csv(uploaded_file)
    elif name.endswith('.xlsx'):
        return _parse_xlsx(uploaded_file)
    else:
        raise FileParseError(
            "Unsupported file type. Please upload a .csv or .xlsx file."
        )


def _normalize_header(header):
    if header is None:
        return ''
    return str(header).strip().lower().replace(' ', '_')


def _parse_csv(uploaded_file):
    raw_bytes = uploaded_file.read()
    try:
        decoded = raw_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        raise FileParseError("Could not read the CSV file. Please save it as UTF-8.")

    reader = csv.DictReader(io.StringIO(decoded))
    if not reader.fieldnames:
        raise FileParseError("The CSV file appears to be empty or has no header row.")

    rows = []
    for raw_row in reader:
        if all((v is None or str(v).strip() == '') for v in raw_row.values()):
            continue  # skip fully blank rows
        row = {}
        for original_key, value in raw_row.items():
            key = _normalize_header(original_key)
            row[key] = value.strip() if isinstance(value, str) else value
        rows.append(row)
    return rows


def _parse_xlsx(uploaded_file):
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as e:
        raise FileParseError(f"Could not read the Excel file: {e}")

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise FileParseError("The Excel file appears to be empty.")

    headers = [_normalize_header(h) for h in header_row]

    rows = []
    for raw_row in rows_iter:
        if all(cell is None or str(cell).strip() == '' for cell in raw_row):
            continue  # skip fully blank rows
        row = {}
        for key, value in zip(headers, raw_row):
            if isinstance(value, str):
                value = value.strip()
            row[key] = value
        rows.append(row)
    return rows