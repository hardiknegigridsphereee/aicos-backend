# finance/views.py
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill

from rest_framework import viewsets, status, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, extend_schema_view

from tenants.views import TenantAwareModelViewSet
from accounts.permissions import IsTeacherOrStaff, IsStudent, IsParent
from profiles.models import StudentProfile, ParentProfile, ParentStudentMapping
from academics.models import StudentEnrollment

from finance.models import FeeStructure, StudentFee, FeeTransaction
from finance.serializers import (
    FeeStructureSerializer, StudentFeeSerializer, FeeTransactionSerializer,
    StudentFeeDetailSerializer
)
from finance.services import FeeBulkUploadService


@extend_schema_view(
    list=extend_schema(summary="List fee structures"),
    create=extend_schema(summary="Create fee structure"),
    retrieve=extend_schema(summary="Get fee structure details"),
    update=extend_schema(summary="Update fee structure"),
    partial_update=extend_schema(summary="Partially update fee structure"),
    destroy=extend_schema(summary="Delete fee structure"),
)
class FeeStructureViewSet(TenantAwareModelViewSet):
    queryset = FeeStructure.objects.select_related('academic_year', 'class_level').all()
    serializer_class = FeeStructureSerializer
    permission_classes = [IsAuthenticated, IsTeacherOrStaff]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['academic_year', 'class_level', 'is_active']
    search_fields = ['class_level__name', 'academic_year__name']


@extend_schema_view(
    list=extend_schema(summary="List student fees"),
    retrieve=extend_schema(summary="Get student fee details"),
)
class StudentFeeViewSet(TenantAwareModelViewSet):
    queryset = StudentFee.objects.select_related(
        'student__user', 'fee_structure__class_level', 'fee_structure__academic_year'
    ).all()
    serializer_class = StudentFeeSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'my_fees', 'my_fee_summary']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsTeacherOrStaff()]

    def get_serializer_class(self):
        if self.action in ['retrieve', 'my_fees']:
            return StudentFeeDetailSerializer
        return StudentFeeSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        
        if hasattr(user, 'studentprofile'):
            student = user.studentprofile
            return qs.filter(student=student)
        
        if hasattr(user, 'parentprofile'):
            parent = user.parentprofile
            student_ids = ParentStudentMapping.objects.filter(
                parent=parent,
                school=user.school
            ).values_list('student_id', flat=True)
            return qs.filter(student_id__in=student_ids)
        
        return qs

    @action(detail=False, methods=['get'], url_path='me')
    def my_fees(self, request):
        try:
            student = StudentProfile.objects.get(user=request.user, school=request.user.school)
        except StudentProfile.DoesNotExist:
            return Response(
                {"detail": "Student profile not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        fees = StudentFee.objects.filter(
            student=student,
            school=request.user.school,
            is_active=True
        ).select_related('fee_structure__class_level', 'fee_structure__academic_year')
        
        serializer = StudentFeeDetailSerializer(fees, many=True)
        return Response({
            "count": fees.count(),
            "results": serializer.data
        })

    @action(detail=False, methods=['get'], url_path='me/summary')
    def my_fee_summary(self, request):
        try:
            student = StudentProfile.objects.get(user=request.user, school=request.user.school)
        except StudentProfile.DoesNotExist:
            return Response(
                {"detail": "Student profile not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        fees = StudentFee.objects.filter(
            student=student,
            school=request.user.school
        )
        
        total_fee = sum(f.total_fee for f in fees)
        total_paid = sum(f.amount_paid for f in fees)
        balance = total_fee - total_paid
        
        return Response({
            "student_name": f"{student.user.first_name} {student.user.last_name}",
            "enrollment_number": student.enrollment_number,
            "total_fee": float(total_fee),
            "total_paid": float(total_paid),
            "balance_due": float(balance),
            "fee_count": fees.count(),
            "pending_fees": fees.filter(status='Pending').count(),
            "overdue_fees": fees.filter(status='Overdue').count(),
            "paid_fees": fees.filter(status='Paid').count()
        })

    @action(detail=False, methods=['post'], url_path='bulk-upload', parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        """
        POST /api/v1/finance/student-fees/bulk-upload/
        Upload Excel file with fee data
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'File must be an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            service = FeeBulkUploadService(
                school=request.user.school,
                user=request.user
            )
            
            options = {
                'update_existing': request.data.get('update_existing', 'true').lower() == 'true',
                'batch_size': int(request.data.get('batch_size', 50)),
            }
            
            results = service.process(
                file_buffer=file.read(),
                file_name=file.name,
                options=options
            )
            
            return Response({
                'summary': {
                    'total': results['total'],
                    'successful': results['successful'],
                    'failed': results['failed'],
                    'created': len(results['created']),
                    'updated': len(results['updated']),
                    'skipped': len(results['skipped']),
                },
                'errors': results['errors'][:20]
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='bulk-export')
    def bulk_export(self, request):
        """
        GET /api/v1/finance/student-fees/bulk-export/
        Exports all student fee records as an .xlsx file, using the exact
        same column schema as the bulk-upload template (student_email,
        enrollment_number, first_name, last_name, class_name, section_name,
        academic_year, tuition_fee, transport_fee, library_fee, lab_fee,
        sports_fee, miscellaneous, due_date, amount_paid), plus a few
        read-only computed columns (total_fee, balance_due, status) at the
        end. Since the first columns match the upload template 1:1, the
        exported file can be edited and re-imported directly via bulk-upload.
        """
        fees = self.get_queryset().select_related(
            'student__user',
            'enrollment__section',
            'fee_structure__class_level',
            'fee_structure__academic_year',
        )

        headers = [
            'student_email', 'enrollment_number', 'first_name', 'last_name',
            'class_name', 'section_name', 'academic_year',
            'tuition_fee', 'transport_fee', 'library_fee', 'lab_fee',
            'sports_fee', 'miscellaneous', 'due_date', 'amount_paid',
            'total_fee', 'balance_due', 'status',
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Fees"

        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for f in fees:
            student = f.student
            user = student.user
            section_name = f.enrollment.section.name if f.enrollment and f.enrollment.section else ""

            ws.append([
                user.email,
                student.enrollment_number,
                user.first_name,
                user.last_name,
                f.fee_structure.class_level.name,
                section_name,
                f.fee_structure.academic_year.name,
                float(f.tuition_fee),
                float(f.transport_fee),
                float(f.library_fee),
                float(f.lab_fee),
                float(f.sports_fee),
                float(f.miscellaneous),
                f.due_date.isoformat() if f.due_date else "",
                float(f.amount_paid),
                float(f.total_fee),
                float(f.balance_due),
                f.status,
            ])

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"student-fees-export-{timezone.now().date().isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class FeeTransactionViewSet(TenantAwareModelViewSet):
    queryset = FeeTransaction.objects.select_related('student__user', 'student_fee').all()
    serializer_class = FeeTransactionSerializer
    permission_classes = [IsAuthenticated, IsTeacherOrStaff]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['student', 'payment_method', 'status']
    search_fields = ['transaction_id', 'reference_number']
    ordering_fields = ['payment_date', 'amount']
    ordering = ['-payment_date']
    